from __future__ import annotations

import argparse
import copy
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:  # pragma: no cover
    Image = None
    ImageDraw = None
    ImageFont = None


PALLET_LARGO_CM = 120
PALLET_ANCHO_CM = 100
PALLET_ALTO_CM = 170
PALLET_BASE_CM2 = PALLET_LARGO_CM * PALLET_ANCHO_CM
PALLET_VOLUMEN_CM3 = PALLET_BASE_CM2 * PALLET_ALTO_CM

STACK_RESISTENTE = 0
STACK_MEDIO = 1
STACK_FRAGIL = 2

RUTA_CATALOGO_ZM040_FIJO = Path(__file__).with_name("zm040_catalogo_fijo.csv")


@dataclass(frozen=True)
class ItemCarga:
    uid: str
    material: str
    descripcion: str
    largo_cm: float
    ancho_cm: float
    alto_cm: float
    peso_kg: float
    stack_rank: int
    stop_id: str
    stop_name: str
    stop_index: int
    ubicacion_almacen: str
    zona_almacen: str
    warehouse_rank: int
    es_retorno: bool
    ruta: str
    transporte: str
    unidad_venta: str
    warehouse_x: float
    warehouse_y: float
    warehouse_area: str
    warehouse_component: str
    warehouse_distance_to_dock: float

    @property
    def volumen_cm3(self) -> float:
        return self.largo_cm * self.ancho_cm * self.alto_cm

    @property
    def base_area_cm2(self) -> float:
        return self.largo_cm * self.ancho_cm

    def orientaciones_base(self) -> List[Tuple[float, float]]:
        opciones = [(self.largo_cm, self.ancho_cm), (self.ancho_cm, self.largo_cm)]
        orientaciones = []
        for largo, ancho in opciones:
            if largo <= PALLET_LARGO_CM and ancho <= PALLET_ANCHO_CM:
                orientaciones.append((largo, ancho))
        return list(dict.fromkeys(orientaciones))


@dataclass
class ColocacionEntrega:
    uid: str
    material: str
    descripcion: str
    slot_id: str
    lado: str
    fila: int
    stop_id: str
    stop_name: str
    stop_index: int
    layer_index: int
    x_cm: float
    y_cm: float
    z_cm: float
    largo_colocado_cm: float
    ancho_colocado_cm: float
    alto_cm: float
    rotado: bool
    peso_kg: float
    volumen_cm3: float
    zona_almacen: str
    warehouse_x: float
    warehouse_y: float
    warehouse_component: str


@dataclass
class PlanRetorno:
    uid: str
    material: str
    descripcion: str
    stop_id: str
    stop_name: str
    stop_index: int
    slot_id: str
    lado: str
    fila: int
    peso_kg: float
    volumen_cm3: float
    estrategia: str


@dataclass
class Estante:
    y_cm: float
    alto_estante_cm: float
    usado_x_cm: float = 0.0


@dataclass
class CapaPalet:
    z_cm: float
    alto_capa_cm: float
    stack_rank_soporte: int
    index: int
    items: List[ColocacionEntrega] = field(default_factory=list)
    estantes: List[Estante] = field(default_factory=list)
    stop_set: set = field(default_factory=set)

    def admite_stop(self, stop_index: int, max_diff: int = 1) -> bool:
        if not self.stop_set:
            return True
        return abs(stop_index - min(self.stop_set)) <= max_diff and abs(stop_index - max(self.stop_set)) <= max_diff

    def intentar_colocar(self, item: ItemCarga, slot_id: str, lado: str, fila: int) -> Optional[ColocacionEntrega]:
        if item.alto_cm > self.alto_capa_cm or item.stack_rank < self.stack_rank_soporte or not self.admite_stop(item.stop_index):
            return None

        for largo, ancho in item.orientaciones_base():
            for estante in self.estantes:
                if ancho <= estante.alto_estante_cm and estante.usado_x_cm + largo <= PALLET_LARGO_CM:
                    colocacion = ColocacionEntrega(
                        uid=item.uid,
                        material=item.material,
                        descripcion=item.descripcion,
                        slot_id=slot_id,
                        lado=lado,
                        fila=fila,
                        stop_id=item.stop_id,
                        stop_name=item.stop_name,
                        stop_index=item.stop_index,
                        layer_index=self.index,
                        x_cm=estante.usado_x_cm,
                        y_cm=estante.y_cm,
                        z_cm=self.z_cm,
                        largo_colocado_cm=largo,
                        ancho_colocado_cm=ancho,
                        alto_cm=item.alto_cm,
                        rotado=(largo, ancho) != (item.largo_cm, item.ancho_cm),
                        peso_kg=item.peso_kg,
                        volumen_cm3=item.volumen_cm3,
                        zona_almacen=item.zona_almacen,
                        warehouse_x=item.warehouse_x,
                        warehouse_y=item.warehouse_y,
                        warehouse_component=item.warehouse_component,
                    )
                    estante.usado_x_cm += largo
                    self.items.append(colocacion)
                    self.stop_set.add(item.stop_index)
                    return colocacion

            usado_y = sum(est.alto_estante_cm for est in self.estantes)
            if usado_y + ancho <= PALLET_ANCHO_CM:
                nuevo_estante = Estante(y_cm=usado_y, alto_estante_cm=ancho, usado_x_cm=largo)
                self.estantes.append(nuevo_estante)
                colocacion = ColocacionEntrega(
                    uid=item.uid,
                    material=item.material,
                    descripcion=item.descripcion,
                    slot_id=slot_id,
                    lado=lado,
                    fila=fila,
                    stop_id=item.stop_id,
                    stop_name=item.stop_name,
                    stop_index=item.stop_index,
                    layer_index=self.index,
                    x_cm=0.0,
                    y_cm=usado_y,
                    z_cm=self.z_cm,
                    largo_colocado_cm=largo,
                    ancho_colocado_cm=ancho,
                    alto_cm=item.alto_cm,
                    rotado=(largo, ancho) != (item.largo_cm, item.ancho_cm),
                    peso_kg=item.peso_kg,
                    volumen_cm3=item.volumen_cm3,
                    zona_almacen=item.zona_almacen,
                    warehouse_x=item.warehouse_x,
                    warehouse_y=item.warehouse_y,
                    warehouse_component=item.warehouse_component,
                )
                self.items.append(colocacion)
                self.stop_set.add(item.stop_index)
                return colocacion

        return None


@dataclass
class SlotCamion:
    slot_id: str
    fila: int
    lado: str
    coord_x: float
    coord_y: float
    alto_max_entrega_cm: float
    reserva_retorno_cm3: float = 0.0
    capas: List[CapaPalet] = field(default_factory=list)
    entregas_por_stop: Dict[int, float] = field(default_factory=lambda: defaultdict(float))
    peso_por_stop: Dict[int, float] = field(default_factory=lambda: defaultdict(float))
    zonas: Counter = field(default_factory=Counter)
    stops: Counter = field(default_factory=Counter)
    volumen_entregado_cm3: float = 0.0
    peso_entregado_kg: float = 0.0
    volumen_retorno_asignado_cm3: float = 0.0
    warehouse_x_sum: float = 0.0
    warehouse_y_sum: float = 0.0
    warehouse_count: int = 0
    warehouse_components: Counter = field(default_factory=Counter)

    def altura_usada_cm(self) -> float:
        if not self.capas:
            return 0.0
        ultima = self.capas[-1]
        return ultima.z_cm + ultima.alto_capa_cm

    def volumen_total_disponible_cm3(self) -> float:
        return PALLET_BASE_CM2 * self.alto_max_entrega_cm

    def capacidad_dinamica_retorno_cm3(self) -> float:
        return self.reserva_retorno_cm3 + sum(self.entregas_por_stop.values())

    def volumen_retorno_libre_cm3(self, entregado_hasta_stop: int) -> float:
        volumen_liberado = sum(vol for stop, vol in self.entregas_por_stop.items() if stop <= entregado_hasta_stop)
        libre = self.reserva_retorno_cm3 + volumen_liberado - self.volumen_retorno_asignado_cm3
        return max(0.0, libre)

    def warehouse_centroid(self) -> Tuple[float, float]:
        if self.warehouse_count <= 0:
            return 0.0, 0.0
        return self.warehouse_x_sum / self.warehouse_count, self.warehouse_y_sum / self.warehouse_count

    def intentar_colocar_entrega(self, item: ItemCarga) -> Optional[ColocacionEntrega]:
        if item.largo_cm > PALLET_LARGO_CM and item.ancho_cm > PALLET_LARGO_CM:
            return None
        if min(item.largo_cm, item.ancho_cm) > PALLET_ANCHO_CM or item.alto_cm > self.alto_max_entrega_cm:
            return None
        if self.volumen_entregado_cm3 + item.volumen_cm3 > self.volumen_total_disponible_cm3():
            return None

        for capa in reversed(self.capas):
            colocacion = capa.intentar_colocar(item, self.slot_id, self.lado, self.fila)
            if colocacion:
                self._registrar_colocacion(item)
                return colocacion

        altura_actual = self.altura_usada_cm()
        if altura_actual + item.alto_cm > self.alto_max_entrega_cm:
            return None
        if self.capas and item.stack_rank < self.capas[-1].stack_rank_soporte:
            return None

        nueva_capa = CapaPalet(
            z_cm=altura_actual,
            alto_capa_cm=item.alto_cm,
            stack_rank_soporte=item.stack_rank,
            index=len(self.capas),
        )
        colocacion = nueva_capa.intentar_colocar(item, self.slot_id, self.lado, self.fila)
        if not colocacion:
            return None

        self.capas.append(nueva_capa)
        self._registrar_colocacion(item)
        return colocacion

    def _registrar_colocacion(self, item: ItemCarga) -> None:
        self.volumen_entregado_cm3 += item.volumen_cm3
        self.peso_entregado_kg += item.peso_kg
        self.entregas_por_stop[item.stop_index] += item.volumen_cm3
        self.peso_por_stop[item.stop_index] += item.peso_kg
        self.zonas[item.zona_almacen] += 1
        self.stops[item.stop_index] += 1
        self.warehouse_x_sum += item.warehouse_x
        self.warehouse_y_sum += item.warehouse_y
        self.warehouse_count += 1
        self.warehouse_components[item.warehouse_component] += 1


@dataclass(frozen=True)
class WarehouseComponent:
    name: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    size: int

    @property
    def center_x(self) -> float:
        return (self.min_col + self.max_col) / 2.0

    @property
    def center_y(self) -> float:
        return (self.min_row + self.max_row) / 2.0


@dataclass(frozen=True)
class WarehouseLayoutModel:
    max_row: int
    max_col: int
    dock_x: float
    dock_y: float
    components: Dict[str, WarehouseComponent]
    cell_values: Dict[Tuple[int, int], int]


def _columna_por_fragmento(df: pd.DataFrame, fragmento: str, preferir_exacta: bool = False) -> str:
    candidatas = [c for c in df.columns if fragmento.lower() in c.lower()]
    if not candidatas:
        raise KeyError(f"No se encontro una columna que contenga '{fragmento}'")
    if preferir_exacta:
        exactas = [c for c in candidatas if c.lower() == fragmento.lower()]
        if exactas:
            return exactas[0]
    return candidatas[0]


def _normalizar_numero(valor: object, default: float = 0.0) -> float:
    if pd.isna(valor):
        return default
    try:
        return float(valor)
    except (TypeError, ValueError):
        return default


def _factor_unidad_a_cm(unidad: object) -> float:
    texto = str(unidad).strip().upper()
    if texto == "MM":
        return 0.1
    if texto in {"M", "MT"}:
        return 100.0
    return 1.0


def _construir_layout_modelo(ruta_layout: Optional[str]) -> Optional[WarehouseLayoutModel]:
    if not ruta_layout or load_workbook is None:
        return None

    ws = load_workbook(ruta_layout, data_only=True)["DDI MOLLET"]
    filled: Dict[Tuple[int, int], int] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            valor = ws.cell(r, c).value
            if valor not in (None, ""):
                try:
                    filled[(r, c)] = int(valor)
                except (TypeError, ValueError):
                    continue

    pendientes = set(filled)
    componentes_raw: List[Tuple[int, int, int, int, int]] = []
    while pendientes:
        inicio = pendientes.pop()
        stack = [inicio]
        componente = [inicio]
        while stack:
            r, c = stack.pop()
            for nr, nc in ((r + 1, c), (r - 1, c), (r, c + 1), (r, c - 1)):
                if (nr, nc) in pendientes:
                    pendientes.remove((nr, nc))
                    stack.append((nr, nc))
                    componente.append((nr, nc))

        rows = [r for r, _ in componente]
        cols = [c for _, c in componente]
        componentes_raw.append((len(componente), min(rows), max(rows), min(cols), max(cols)))

    componentes_raw.sort(reverse=True)
    nombres = [
        "main_storage",
        "compact_storage",
        "special_floor",
        "blue_lane_long",
        "upper_blue",
        "upper_green_low",
        "upper_green_mid",
        "upper_green_high",
        "upper_green_top",
    ]
    componentes: Dict[str, WarehouseComponent] = {}
    for idx, comp in enumerate(componentes_raw):
        size, min_row, max_row, min_col, max_col = comp
        nombre = nombres[idx] if idx < len(nombres) else f"component_{idx}"
        componentes[nombre] = WarehouseComponent(
            name=nombre,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            size=size,
        )

    main = componentes.get("main_storage")
    dock_x = (main.min_col - 2.0) if main else 1.0
    dock_y = (main.min_row - 2.0) if main else 1.0
    return WarehouseLayoutModel(
        max_row=ws.max_row,
        max_col=ws.max_column,
        dock_x=dock_x,
        dock_y=dock_y,
        components=componentes,
        cell_values=filled,
    )


def _coord_centro_component(modelo: Optional[WarehouseLayoutModel], nombre: str, fallback: Tuple[float, float]) -> Tuple[float, float]:
    if not modelo:
        return fallback
    comp = modelo.components.get(nombre)
    if not comp:
        return fallback
    return comp.center_x, comp.center_y


def _coord_desde_ubicacion(ubicacion: str, modelo: Optional[WarehouseLayoutModel]) -> Tuple[float, float, str, str, float]:
    texto = str(ubicacion or "").strip().upper()
    if not texto:
        return 0.0, 0.0, "SIN_UBIC", "unknown", 0.0

    if not modelo:
        zona, pasillo = _parsear_ubicacion(texto)
        x = max(1.0, float(sum(ord(ch) for ch in zona) % 50))
        y = max(1.0, float(pasillo if pasillo != 9999 else 1))
        return x, y, zona, "fallback", x + y

    dock_x, dock_y = modelo.dock_x, modelo.dock_y
    main = modelo.components.get("main_storage")
    compact = modelo.components.get("compact_storage")
    special = modelo.components.get("special_floor")
    upper_blue = modelo.components.get("upper_blue")

    if texto in {"ZCG", "A0DISTRIDA"}:
        x = modelo.max_col + (8.0 if texto == "ZCG" else 14.0)
        y = (special.center_y if special else modelo.max_row / 2.0)
        return x, y, texto[:2], "external", abs(x - dock_x) + abs(y - dock_y)
    if texto == "CAMARA":
        x, y = _coord_centro_component(modelo, "upper_blue", (modelo.max_col * 0.8, modelo.max_row * 0.1))
        return x, y, "CA", "cold_room", abs(x - dock_x) + abs(y - dock_y)
    if texto in {"ENVASE", "PLV", "AAAAAA"} or texto.startswith("PL"):
        x, y = _coord_centro_component(modelo, "special_floor", (modelo.max_col * 0.8, modelo.max_row * 0.85))
        return x, y, texto[:2], "floor_special", abs(x - dock_x) + abs(y - dock_y)
    if texto.startswith("EN"):
        x, y = _coord_centro_component(modelo, "special_floor", (modelo.max_col * 0.75, modelo.max_row * 0.25))
        return x, y, texto[:2], "floor_special", abs(x - dock_x) + abs(y - dock_y)

    if main:
        prefix_match = re.match(r"^(?P<prefix>[A-Z]{2})(?P<rest>.*)$", texto)
        if prefix_match:
            prefix = prefix_match.group("prefix")
            resto = prefix_match.group("rest")
            group_major = min(5, max(0, ord(prefix[0]) - ord("A")))
            group_minor = min(2, max(0, ord(prefix[1]) - ord("A"))) if prefix[1].isalpha() else 0
            digits = re.findall(r"\d{1,3}", resto)
            aisle = int(digits[0]) if digits else 1
            level = int(digits[-1][-1]) if digits else 1
            suffix_letter = next((ch for ch in resto if ch.isalpha()), "A")

            usable_x = max(6.0, (main.max_col - main.min_col - 4))
            usable_y = max(10.0, (main.max_row - main.min_row - 6))
            max_aisle = 12.0
            base_x = main.min_col + 2.0 + (group_major + 0.5) * (usable_x / 6.0)
            base_x += group_minor * (usable_x / 24.0)
            base_x += {"A": -0.7, "B": 0.0, "C": 0.7}.get(suffix_letter, 0.0)
            base_y = main.max_row - 2.0 - (min(aisle, max_aisle) - 0.5) * (usable_y / max_aisle)
            base_y -= min(level, 9) * 0.2
            return base_x, base_y, prefix, "main_storage", abs(base_x - dock_x) + abs(base_y - dock_y)

    if compact:
        x, y = compact.center_x, compact.center_y
        return x, y, texto[:2], "compact_storage", abs(x - dock_x) + abs(y - dock_y)
    if upper_blue:
        x, y = upper_blue.center_x, upper_blue.center_y
        return x, y, texto[:2], "upper_blue", abs(x - dock_x) + abs(y - dock_y)

    return 0.0, 0.0, texto[:2], "unknown", 0.0


def _clasificar_stack_rank(descripcion: str, unidad_venta: str, es_retorno: bool) -> int:
    texto = f"{descripcion} {unidad_venta}".upper()
    if es_retorno:
        return STACK_MEDIO
    if any(token in texto for token in ("BARRIL", "KEG", "BIDON", "METAL")):
        return STACK_RESISTENTE
    if any(token in texto for token in ("CAJA", "C.C.", "CARTON", "CJ", "CRATE")):
        return STACK_MEDIO
    if any(token in texto for token in ("PET", "PLAST", "BOTELLA", "VASO")):
        return STACK_FRAGIL
    return STACK_MEDIO


def _es_linea_retorno(material: str, descripcion: str) -> bool:
    texto = descripcion.upper()
    return material.startswith("3ENV") or any(token in texto for token in ("VACIO", "VACI", "ENVASE", "RETORNO"))


def _parsear_ubicacion(ubicacion: str) -> Tuple[str, int]:
    if not ubicacion:
        return "ZZ", 9999
    texto = str(ubicacion).upper()
    zona = texto[:2]
    numeros = re.findall(r"\d+", texto)
    pasillo = int(numeros[0]) if numeros else 9999
    return zona, pasillo


def _estimar_dimensiones(descripcion: str, unidad_venta: str, es_retorno: bool) -> Tuple[float, float, float]:
    texto = f"{descripcion} {unidad_venta}".upper()
    if es_retorno:
        if "1/5" in texto:
            return 60.0, 40.0, 30.0
        if "1/3" in texto:
            return 60.0, 40.0, 26.0
        return 60.0, 40.0, 25.0
    if any(token in texto for token in ("BARRIL", "KEG", "BIDON")):
        return 40.0, 40.0, 60.0
    if any(token in texto for token in ("CAJ", "CAJA", "C.C.")):
        return 40.0, 30.0, 25.0
    if any(token in texto for token in ("SACO", "BOLSA")):
        return 45.0, 35.0, 15.0
    return 35.0, 25.0, 20.0


def _estimar_peso_kg(descripcion: str, unidad_venta: str, volumen_cm3: float, es_retorno: bool) -> float:
    texto = f"{descripcion} {unidad_venta}".upper()
    if es_retorno:
        return 2.5
    if any(token in texto for token in ("BARRIL", "KEG", "BIDON")):
        return 28.0
    if any(token in texto for token in ("CAJ", "CAJA", "C.C.")):
        return min(18.0, max(4.0, volumen_cm3 / 4500.0))
    return min(12.0, max(1.0, volumen_cm3 / 7000.0))


def _preparar_catalogo_dimensiones(df_zm040: pd.DataFrame) -> Tuple[Dict[Tuple[str, str], dict], Dict[str, List[dict]]]:
    if {"material", "uma", "largo_cm", "ancho_cm", "alto_cm", "peso_kg"}.issubset(df_zm040.columns):
        base = df_zm040.copy()
        base["material"] = base["material"].astype(str).str.strip()
        base["uma"] = base["uma"].astype(str).str.strip()
        for col in ("largo_cm", "ancho_cm", "alto_cm", "peso_kg"):
            base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0.0)

        registros_por_clave: Dict[Tuple[str, str], dict] = {}
        registros_por_material: Dict[str, List[dict]] = defaultdict(list)
        for _, fila in base.iterrows():
            material = str(fila["material"]).strip()
            uma = str(fila["uma"]).strip()
            if not material:
                continue
            registro = {
                "material": material,
                "uma": uma,
                "largo_cm": float(fila["largo_cm"]),
                "ancho_cm": float(fila["ancho_cm"]),
                "alto_cm": float(fila["alto_cm"]),
                "peso_kg": float(fila["peso_kg"]),
            }
            registros_por_clave[(material, uma)] = registro
            registros_por_material[material].append(registro)
        return registros_por_clave, registros_por_material

    zm = df_zm040.copy()
    columnas_unidad = [c for c in zm.columns if str(c).startswith("Unidad")]
    col_u_largo = columnas_unidad[0] if len(columnas_unidad) > 0 else None
    col_u_ancho = columnas_unidad[1] if len(columnas_unidad) > 1 else None
    col_u_alto = columnas_unidad[2] if len(columnas_unidad) > 2 else None

    for col in ("Longitud", "Ancho", "Altura", "Peso bruto", "Peso neto", "Volumen"):
        if col in zm.columns:
            zm[col] = pd.to_numeric(zm[col], errors="coerce")

    zm = zm.fillna({"UMA": "", "Material": ""})
    registros_por_clave: Dict[Tuple[str, str], dict] = {}
    registros_por_material: Dict[str, List[dict]] = defaultdict(list)

    for _, fila in zm.iterrows():
        material = str(fila["Material"]).strip()
        uma = str(fila["UMA"]).strip()
        largo_raw = _normalizar_numero(fila.get("Longitud"), default=0.0) or 0.0
        ancho_raw = _normalizar_numero(fila.get("Ancho"), default=0.0) or 0.0
        alto_raw = _normalizar_numero(fila.get("Altura"), default=0.0) or 0.0
        largo = largo_raw * (_factor_unidad_a_cm(fila.get(col_u_largo)) if col_u_largo else 1.0)
        ancho = ancho_raw * (_factor_unidad_a_cm(fila.get(col_u_ancho)) if col_u_ancho else 1.0)
        alto = alto_raw * (_factor_unidad_a_cm(fila.get(col_u_alto)) if col_u_alto else 1.0)
        peso_bruto = _normalizar_numero(fila.get("Peso bruto"))
        peso_neto = _normalizar_numero(fila.get("Peso neto"))

        if not material:
            continue

        registro = {
            "material": material,
            "uma": uma,
            "largo_cm": largo,
            "ancho_cm": ancho,
            "alto_cm": alto,
            "peso_kg": peso_bruto or peso_neto,
        }
        registros_por_clave[(material, uma)] = registro
        registros_por_material[material].append(registro)

    return registros_por_clave, registros_por_material


def _cargar_catalogo_zm040(ruta_zm040: Optional[str]) -> pd.DataFrame:
    ruta_preferida = Path(ruta_zm040) if ruta_zm040 else RUTA_CATALOGO_ZM040_FIJO
    if not ruta_preferida.exists():
        raise FileNotFoundError(f"No se encontro el catalogo ZM040 en {ruta_preferida}")

    if ruta_preferida.suffix.lower() == ".csv":
        return pd.read_csv(ruta_preferida)
    return pd.read_excel(ruta_preferida)


def _buscar_dimensiones_y_peso(
    material: str,
    unidad_venta: str,
    descripcion: str,
    es_retorno: bool,
    catalogo_exacto: Dict[Tuple[str, str], dict],
    catalogo_material: Dict[str, List[dict]],
) -> Tuple[float, float, float, float]:
    clave = (material, unidad_venta)
    if clave in catalogo_exacto:
        reg = catalogo_exacto[clave]
        if reg["largo_cm"] > 0 and reg["ancho_cm"] > 0 and reg["alto_cm"] > 0:
            peso = reg["peso_kg"] if reg["peso_kg"] > 0 else 0.0
            if peso <= 0:
                peso = _estimar_peso_kg(descripcion, unidad_venta, reg["largo_cm"] * reg["ancho_cm"] * reg["alto_cm"], es_retorno)
            return reg["largo_cm"], reg["ancho_cm"], reg["alto_cm"], peso

    candidatos = []
    for reg in catalogo_material.get(material, []):
        if reg["uma"] == "PAL":
            continue
        if reg["largo_cm"] > 0 and reg["ancho_cm"] > 0 and reg["alto_cm"] > 0:
            volumen = reg["largo_cm"] * reg["ancho_cm"] * reg["alto_cm"]
            candidatos.append((volumen, reg))
    if candidatos:
        _, reg = min(candidatos, key=lambda x: x[0])
        peso = reg["peso_kg"] if reg["peso_kg"] > 0 else 0.0
        if peso <= 0:
            peso = _estimar_peso_kg(descripcion, unidad_venta, reg["largo_cm"] * reg["ancho_cm"] * reg["alto_cm"], es_retorno)
        return reg["largo_cm"], reg["ancho_cm"], reg["alto_cm"], peso

    largo, ancho, alto = _estimar_dimensiones(descripcion, unidad_venta, es_retorno)
    peso = _estimar_peso_kg(descripcion, unidad_venta, largo * ancho * alto, es_retorno)
    return largo, ancho, alto, peso


def cargar_items_desde_excels(
    ruta_hackaton: str,
    ruta_zm040: Optional[str] = None,
    *,
    ruta_layout: Optional[str] = "Layout Mollet.xlsx",
    ruta: Optional[str] = None,
    transporte: Optional[str] = None,
    stop_sequence: Optional[Sequence[str]] = None,
    stop_ids: Optional[Sequence[str]] = None,
    max_stops: Optional[int] = None,
    usar_columna_stop: str = "Nombre 1",
) -> Tuple[List[ItemCarga], Dict[int, str], Dict[str, int]]:
    if not ruta and not transporte:
        raise ValueError("Debes indicar una ruta o un transporte para construir un camion concreto.")

    df_detalle = pd.read_excel(ruta_hackaton, sheet_name="Detalle entrega")
    df_materiales = pd.read_excel(ruta_hackaton, sheet_name="Materiales zubic")
    df_zm040 = _cargar_catalogo_zm040(ruta_zm040)

    if ruta:
        df_detalle = df_detalle[df_detalle["Ruta"].astype(str) == str(ruta)].copy()
    if transporte:
        df_detalle = df_detalle[df_detalle["Transporte"].astype(str) == str(transporte)].copy()

    if df_detalle.empty:
        raise ValueError("No hay lineas para el filtro indicado.")

    col_material = "Material"
    col_desc = _columna_por_fragmento(df_detalle, "Denomin")
    col_qty = _columna_por_fragmento(df_detalle, "Cantidad")
    col_unidad = _columna_por_fragmento(df_detalle, "Un.medida")
    col_stop_id = [c for c in df_detalle.columns if c.startswith("Destinatario") and ".1" in c][0]
    col_stop_name = usar_columna_stop if usar_columna_stop in df_detalle.columns else col_stop_id

    if stop_ids:
        permitidos = {str(stop) for stop in stop_ids}
        df_detalle = df_detalle[df_detalle[col_stop_id].astype(str).isin(permitidos)].copy()
    elif max_stops:
        stops_en_orden = pd.Series(df_detalle[col_stop_id].astype(str).tolist()).drop_duplicates().tolist()
        permitidos = set(stops_en_orden[:max_stops])
        df_detalle = df_detalle[df_detalle[col_stop_id].astype(str).isin(permitidos)].copy()

    if df_detalle.empty:
        raise ValueError("Tras aplicar el filtro de paradas no quedan lineas.")

    df_base = df_detalle.merge(df_materiales[[col_material, "Ubic."]], on=col_material, how="left")
    catalogo_exacto, catalogo_material = _preparar_catalogo_dimensiones(df_zm040)
    warehouse_layout = _construir_layout_modelo(ruta_layout)

    if stop_sequence:
        mapa_paradas = {str(stop): idx + 1 for idx, stop in enumerate(stop_sequence)}
    else:
        stops_en_orden = pd.Series(df_base[col_stop_id].astype(str).tolist()).drop_duplicates().tolist()
        mapa_paradas = {str(stop): idx + 1 for idx, stop in enumerate(stops_en_orden)}

    nombres_parada = {
        idx: str(df_base.loc[df_base[col_stop_id].astype(str) == stop, col_stop_name].iloc[0])
        for stop, idx in mapa_paradas.items()
    }

    items: List[ItemCarga] = []
    for _, fila in df_base.iterrows():
        material = str(fila[col_material]).strip()
        descripcion = str(fila[col_desc]).strip()
        unidad_venta = str(fila[col_unidad]).strip()
        stop_id = str(fila[col_stop_id]).strip()
        stop_name = str(fila[col_stop_name]).strip()
        qty = max(1, int(_normalizar_numero(fila[col_qty], default=1.0)))
        ruta_valor = str(fila["Ruta"])
        transporte_valor = str(fila["Transporte"])
        ubicacion = str(fila.get("Ubic.", "")).strip()
        es_retorno = _es_linea_retorno(material, descripcion)
        largo, ancho, alto, peso = _buscar_dimensiones_y_peso(
            material,
            unidad_venta,
            descripcion,
            es_retorno,
            catalogo_exacto,
            catalogo_material,
        )
        zona, pasillo = _parsear_ubicacion(ubicacion)
        warehouse_x, warehouse_y, warehouse_area, warehouse_component, warehouse_distance_to_dock = _coord_desde_ubicacion(
            ubicacion,
            warehouse_layout,
        )
        warehouse_rank = int(warehouse_distance_to_dock * 1000) + (pasillo if pasillo != 9999 else 999)
        stack_rank = _clasificar_stack_rank(descripcion, unidad_venta, es_retorno)
        stop_index = mapa_paradas.setdefault(stop_id, len(mapa_paradas) + 1)

        for unidad in range(qty):
            items.append(
                ItemCarga(
                    uid=f"{transporte_valor}-{material}-{stop_index}-{unidad}-{len(items)}",
                    material=material,
                    descripcion=descripcion,
                    largo_cm=largo,
                    ancho_cm=ancho,
                    alto_cm=alto,
                    peso_kg=peso,
                    stack_rank=stack_rank,
                    stop_id=stop_id,
                    stop_name=stop_name,
                    stop_index=stop_index,
                    ubicacion_almacen=ubicacion,
                    zona_almacen=zona,
                    warehouse_rank=warehouse_rank,
                    es_retorno=es_retorno,
                    ruta=ruta_valor,
                    transporte=transporte_valor,
                    unidad_venta=unidad_venta,
                    warehouse_x=warehouse_x,
                    warehouse_y=warehouse_y,
                    warehouse_area=warehouse_area,
                    warehouse_component=warehouse_component,
                    warehouse_distance_to_dock=warehouse_distance_to_dock,
                )
            )

    nombres_parada = {idx: nombres_parada.get(idx, f"Parada {idx}") for idx in sorted(nombres_parada)}
    return items, nombres_parada, mapa_paradas


def _construir_slots_camion(num_palets: int, reserva_retorno_total_cm3: float) -> List[SlotCamion]:
    if num_palets not in (3, 6, 8):
        raise ValueError("Solo se soportan camiones de 3, 6 u 8 palets.")

    if num_palets == 3:
        layout = [("C", 0), ("C", 1), ("C", 2)]
    elif num_palets == 6:
        layout = [("L", 0), ("R", 0), ("L", 1), ("R", 1), ("L", 2), ("R", 2)]
    else:
        layout = [("L", 0), ("R", 0), ("L", 1), ("R", 1), ("L", 2), ("R", 2), ("L", 3), ("R", 3)]

    slots = []
    filas_totales = max(fila for _, fila in layout) + 1
    slots_ordenados_reserva = sorted(layout, key=lambda t: (t[1], 0 if t[0] == "C" else 1))

    reserva_restante = reserva_retorno_total_cm3
    reservas_por_slot: Dict[Tuple[str, int], float] = {}
    capacidad_slot = PALLET_VOLUMEN_CM3
    for idx, key in enumerate(slots_ordenados_reserva):
        slots_restantes = len(slots_ordenados_reserva) - idx
        cuota = min(capacidad_slot * 0.35, reserva_restante / max(1, slots_restantes))
        reservas_por_slot[key] = max(0.0, cuota)
        reserva_restante -= cuota

    for lado, fila in layout:
        coord_x = 0.0 if lado == "C" else (-1.0 if lado == "L" else 1.0)
        coord_y = float(fila)
        reserva_slot = reservas_por_slot.get((lado, fila), 0.0)
        reserva_altura_cm = math.floor(reserva_slot / PALLET_BASE_CM2)
        alto_max_entrega = max(30.0, PALLET_ALTO_CM - reserva_altura_cm)
        slots.append(
            SlotCamion(
                slot_id=f"{lado}{fila + 1}",
                fila=fila,
                lado=lado,
                coord_x=coord_x,
                coord_y=coord_y,
                alto_max_entrega_cm=alto_max_entrega,
                reserva_retorno_cm3=reserva_slot,
            )
        )

    return slots


def _fila_objetivo_por_parada(stop_index: int, total_paradas: int, filas_totales: int) -> float:
    if total_paradas <= 1 or filas_totales <= 1:
        return 0.0
    proporcion = (stop_index - 1) / (total_paradas - 1)
    return proporcion * (filas_totales - 1)


def _score_slot_entrega(
    slot: SlotCamion,
    item: ItemCarga,
    slots_referencia: List[SlotCamion],
    pesos_objetivo: Dict[str, float],
    total_paradas: int,
) -> float:
    filas_totales = max(s.fila for s in slots_referencia) + 1
    fila_objetivo = _fila_objetivo_por_parada(item.stop_index, total_paradas, filas_totales)
    access_penalty = abs(slot.fila - fila_objetivo)

    if slot.warehouse_count > 0:
        centro_x, centro_y = slot.warehouse_centroid()
        warehouse_penalty = (abs(item.warehouse_x - centro_x) + abs(item.warehouse_y - centro_y)) / 12.0
        if slot.warehouse_components[item.warehouse_component] == 0:
            warehouse_penalty += 1.0
    else:
        warehouse_penalty = item.warehouse_distance_to_dock / 18.0

    stop_span_penalty = 0.0
    if slot.stops:
        min_stop = min(slot.stops)
        max_stop = max(slot.stops)
        if item.stop_index < min_stop:
            stop_span_penalty = min_stop - item.stop_index
        elif item.stop_index > max_stop:
            stop_span_penalty = item.stop_index - max_stop

    volumen_relativo = (slot.volumen_entregado_cm3 + item.volumen_cm3) / max(1.0, slot.volumen_total_disponible_cm3())
    fill_penalty = max(0.0, volumen_relativo - 0.88) * 8.0

    lado_izq = sum(s.peso_entregado_kg for s in slots_referencia if s.lado == "L")
    lado_der = sum(s.peso_entregado_kg for s in slots_referencia if s.lado == "R")
    lado_centro = sum(s.peso_entregado_kg for s in slots_referencia if s.lado == "C")
    if slot.lado == "L":
        lado_izq += item.peso_kg
    elif slot.lado == "R":
        lado_der += item.peso_kg
    else:
        lado_centro += item.peso_kg

    side_penalty = abs(lado_izq - lado_der) / max(1.0, lado_izq + lado_der + lado_centro)

    pesos_totales = [(s.fila, s.peso_entregado_kg + (item.peso_kg if s.slot_id == slot.slot_id else 0.0)) for s in slots_referencia]
    total_peso = sum(p for _, p in pesos_totales)
    centro_long = sum(fila * peso for fila, peso in pesos_totales) / max(1.0, total_peso)
    centro_ideal = (filas_totales - 1) / 2
    long_penalty = abs(centro_long - centro_ideal) / max(1.0, centro_ideal if centro_ideal else 1.0)

    return (
        pesos_objetivo["acceso"] * access_penalty
        + pesos_objetivo["almacen"] * warehouse_penalty
        + pesos_objetivo["espacio"] * fill_penalty
        + pesos_objetivo["mezcla_paradas"] * stop_span_penalty
        + pesos_objetivo["peso_lateral"] * side_penalty
        + pesos_objetivo["peso_longitudinal"] * long_penalty
    )


def _score_slot_retorno(slot: SlotCamion, item: ItemCarga, stop_index: int, slots: List[SlotCamion]) -> float:
    libre = slot.volumen_retorno_libre_cm3(stop_index)
    if libre < item.volumen_cm3:
        return math.inf

    filas_totales = max(s.fila for s in slots) + 1
    access_penalty = slot.fila / max(1.0, filas_totales - 1)
    free_bonus = 1.0 - min(1.0, libre / max(item.volumen_cm3, 1.0))

    lado_izq = sum(s.peso_entregado_kg for s in slots if s.lado == "L")
    lado_der = sum(s.peso_entregado_kg for s in slots if s.lado == "R")
    if slot.lado == "L":
        lado_izq += item.peso_kg
    elif slot.lado == "R":
        lado_der += item.peso_kg
    side_penalty = abs(lado_izq - lado_der) / max(1.0, lado_izq + lado_der)

    return (3.0 * access_penalty) + (1.5 * free_bonus) + side_penalty


def _pico_neto_retorno(items: Iterable[ItemCarga], total_paradas: int) -> float:
    entregas = defaultdict(float)
    retornos = defaultdict(float)
    for item in items:
        if item.es_retorno:
            retornos[item.stop_index] += item.volumen_cm3
        else:
            entregas[item.stop_index] += item.volumen_cm3

    neto = 0.0
    pico = 0.0
    for stop_index in range(1, total_paradas + 1):
        neto += retornos[stop_index] - entregas[stop_index]
        pico = max(pico, neto)
    return max(0.0, pico)


def _construir_picking_por_ciclos(items: Sequence[ItemCarga], df_colocaciones: pd.DataFrame) -> pd.DataFrame:
    if df_colocaciones.empty:
        return pd.DataFrame()

    item_map = {item.uid: item for item in items if not item.es_retorno}
    base = df_colocaciones.sort_values(["layer_index", "fila", "lado", "y_cm", "x_cm"]).copy()
    registros = []

    for layer_index in sorted(base["layer_index"].unique().tolist()):
        candidatos = []
        layer_df = base[base["layer_index"] == layer_index]
        for _, fila in layer_df.iterrows():
            item = item_map.get(fila["uid"])
            if not item:
                continue
            candidatos.append(
                {
                    "uid": item.uid,
                    "material": item.material,
                    "descripcion": item.descripcion,
                    "slot_id": fila["slot_id"],
                    "lado": fila["lado"],
                    "fila_camion": fila["fila"],
                    "cycle": int(layer_index) + 1,
                    "layer_index": int(layer_index),
                    "stop_index": item.stop_index,
                    "stop_name": item.stop_name,
                    "fragilidad_rank": item.stack_rank,
                    "ubicacion_almacen": item.ubicacion_almacen,
                    "warehouse_x": item.warehouse_x,
                    "warehouse_y": item.warehouse_y,
                    "warehouse_component": item.warehouse_component,
                }
            )

        prev_x = 0.0
        prev_y = 0.0
        orden = 1
        pendientes = candidatos[:]
        while pendientes:
            idx_mejor, mejor = min(
                enumerate(pendientes),
                key=lambda t: (
                    abs(t[1]["warehouse_x"] - prev_x) + abs(t[1]["warehouse_y"] - prev_y),
                    t[1]["fragilidad_rank"],
                    t[1]["stop_index"],
                ),
            )
            distancia = abs(mejor["warehouse_x"] - prev_x) + abs(mejor["warehouse_y"] - prev_y)
            mejor["orden_en_ciclo"] = orden
            mejor["distancia_desde_anterior"] = distancia
            registros.append(mejor)
            prev_x = mejor["warehouse_x"]
            prev_y = mejor["warehouse_y"]
            orden += 1
            pendientes.pop(idx_mejor)

    return pd.DataFrame(registros).sort_values(["cycle", "orden_en_ciclo", "slot_id"])


def _construir_palets_finales(items: Sequence[ItemCarga], df_colocaciones: pd.DataFrame) -> pd.DataFrame:
    if df_colocaciones.empty:
        return pd.DataFrame()

    item_map = {item.uid: item for item in items if not item.es_retorno}
    base = df_colocaciones.sort_values(["slot_id", "layer_index", "z_cm", "y_cm", "x_cm"]).copy()
    registros = []
    ordenes_por_slot: Dict[str, int] = defaultdict(int)

    for _, fila in base.iterrows():
        item = item_map.get(fila["uid"])
        if not item:
            continue
        ordenes_por_slot[fila["slot_id"]] += 1
        registros.append(
            {
                "slot_id": fila["slot_id"],
                "lado": fila["lado"],
                "fila_camion": fila["fila"],
                "cycle": int(fila["layer_index"]) + 1,
                "layer_index": int(fila["layer_index"]),
                "orden_en_palet": ordenes_por_slot[fila["slot_id"]],
                "uid": item.uid,
                "material": item.material,
                "descripcion": item.descripcion,
                "stop_index": item.stop_index,
                "stop_name": item.stop_name,
                "fragilidad_rank": item.stack_rank,
                "ubicacion_almacen": item.ubicacion_almacen,
                "x_cm": fila["x_cm"],
                "y_cm": fila["y_cm"],
                "z_cm": fila["z_cm"],
                "largo_colocado_cm": fila["largo_colocado_cm"],
                "ancho_colocado_cm": fila["ancho_colocado_cm"],
                "alto_cm": fila["alto_cm"],
            }
        )

    return pd.DataFrame(registros)


def calcular_factor_estiba_viaje(items: Sequence[ItemCarga], num_palets: int) -> pd.DataFrame:
    entregas = [item for item in items if not item.es_retorno]
    retornos = [item for item in items if item.es_retorno]
    vol_entregas = sum(item.volumen_cm3 for item in entregas)
    vol_retornos = sum(item.volumen_cm3 for item in retornos)
    vol_resistente = sum(item.volumen_cm3 for item in entregas if item.stack_rank == STACK_RESISTENTE)
    vol_fragil = sum(item.volumen_cm3 for item in entregas if item.stack_rank == STACK_FRAGIL)
    ratio_resistente = vol_resistente / max(1.0, vol_entregas)
    ratio_fragil = vol_fragil / max(1.0, vol_entregas)
    pico_retorno = _pico_neto_retorno(items, max((item.stop_index for item in items), default=0))

    if ratio_resistente >= 0.45:
        utilizacion_recomendada = 0.76
        motivo = "mucho_formato_rigido_o_barril"
    elif ratio_fragil >= 0.45:
        utilizacion_recomendada = 0.90
        motivo = "predominio_caja_ajustable"
    else:
        utilizacion_recomendada = 0.84
        motivo = "mezcla_estandar"

    volumen_nominal = num_palets * PALLET_VOLUMEN_CM3
    volumen_util_recomendado = volumen_nominal * utilizacion_recomendada
    volumen_entregable_recomendado = max(0.0, volumen_util_recomendado - pico_retorno)
    cabe_entregas = vol_entregas <= volumen_entregable_recomendado

    return pd.DataFrame(
        [
            {
                "palets": num_palets,
                "volumen_nominal_cm3": volumen_nominal,
                "factor_estiba_recomendado": utilizacion_recomendada,
                "margen_extra_pct": 1.0 - utilizacion_recomendada,
                "volumen_util_recomendado_cm3": volumen_util_recomendado,
                "pico_retorno_cm3": pico_retorno,
                "volumen_entregable_recomendado_cm3": volumen_entregable_recomendado,
                "volumen_entregas_cm3": vol_entregas,
                "volumen_retornos_total_cm3": vol_retornos,
                "ratio_resistente": ratio_resistente,
                "ratio_fragil": ratio_fragil,
                "cabe_según_factor_estiba": cabe_entregas,
                "motivo_factor": motivo,
            }
        ]
    )


def renderizar_plan_camion(resultado: Dict[str, pd.DataFrame], ruta_salida: str) -> Optional[str]:
    if Image is None or resultado["colocaciones_entrega"].empty:
        return None

    df = resultado["colocaciones_entrega"].copy()
    resumen = resultado["resumen_slots"].copy()
    slot_ids = resumen["slot_id"].tolist()
    colores = [
        (82, 121, 255),
        (107, 203, 119),
        (255, 183, 77),
        (240, 98, 146),
        (38, 198, 218),
        (171, 71, 188),
        (255, 112, 67),
        (124, 179, 66),
    ]
    stops = sorted(df["stop_index"].unique().tolist())
    color_by_stop = {stop: colores[idx % len(colores)] for idx, stop in enumerate(stops)}

    slots_por_fila = max(1, len(set(resumen["fila"])))
    scale = 2
    panel_w = int(PALLET_LARGO_CM * scale) + 60
    layer_h = int(PALLET_ANCHO_CM * scale) + 50
    margin = 30
    max_layers = int(df.groupby("slot_id")["layer_index"].max().max()) + 1
    img_w = panel_w * len(slot_ids) + margin * 2
    img_h = margin * 2 + max_layers * layer_h + 70
    img = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()

    draw.text((margin, 10), "Plan de carga por slot y capa", fill="black", font=font)
    for s_idx, slot_id in enumerate(slot_ids):
        slot_df = df[df["slot_id"] == slot_id]
        panel_x = margin + s_idx * panel_w
        draw.text((panel_x, 28), slot_id, fill="black", font=font)
        for layer_index in sorted(slot_df["layer_index"].unique().tolist()):
            layer_df = slot_df[slot_df["layer_index"] == layer_index]
            x0 = panel_x
            y0 = 50 + layer_index * layer_h
            x1 = x0 + int(PALLET_LARGO_CM * scale)
            y1 = y0 + int(PALLET_ANCHO_CM * scale)
            draw.rectangle([x0, y0, x1, y1], outline=(90, 90, 90), width=2)
            draw.text((x0, y1 + 4), f"capa {layer_index}", fill=(70, 70, 70), font=font)
            for _, row in layer_df.iterrows():
                rx0 = x0 + int(row["x_cm"] * scale)
                ry0 = y0 + int(row["y_cm"] * scale)
                rx1 = rx0 + int(row["largo_colocado_cm"] * scale)
                ry1 = ry0 + int(row["ancho_colocado_cm"] * scale)
                color = color_by_stop.get(int(row["stop_index"]), (180, 180, 180))
                draw.rectangle([rx0, ry0, rx1, ry1], fill=color, outline=(30, 30, 30), width=1)
                draw.text((rx0 + 2, ry0 + 2), str(int(row["stop_index"])), fill="black", font=font)

    Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)
    img.save(ruta_salida)
    return str(Path(ruta_salida).resolve())


def renderizar_layout_picking(
    items: Sequence[ItemCarga],
    ruta_layout: Optional[str],
    resultado: Dict[str, pd.DataFrame],
    ruta_salida: str,
) -> Optional[str]:
    if Image is None:
        return None

    modelo = _construir_layout_modelo(ruta_layout)
    if not modelo:
        return None

    ruta_picking = resultado.get("picking_por_ciclos")
    if ruta_picking is None or ruta_picking.empty:
        ruta_picking = _construir_picking_por_ciclos(items, resultado["colocaciones_entrega"])
    scale = 10
    margin = 30
    img_w = int(modelo.max_col * scale + margin * 2 + 220)
    img_h = int(modelo.max_row * scale + margin * 2)
    img = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()
    color_por_valor = {
        1: (225, 225, 225),
        2: (255, 181, 115),
        3: (146, 208, 80),
        4: (92, 141, 255),
        9: (255, 230, 40),
    }
    for (r, c), valor in modelo.cell_values.items():
        color = color_por_valor.get(valor, (235, 235, 235))
        x0 = margin + (c - 1) * scale
        y0 = margin + (r - 1) * scale
        draw.rectangle([x0, y0, x0 + scale - 1, y0 + scale - 1], fill=color, outline=(255, 255, 255))

    dock_px = margin + modelo.dock_x * scale
    dock_py = margin + modelo.dock_y * scale
    draw.ellipse([dock_px - 5, dock_py - 5, dock_px + 5, dock_py + 5], fill=(255, 0, 0))
    draw.text((dock_px + 8, dock_py - 8), "dock", fill="black", font=font)

    if not ruta_picking.empty:
        colores_ciclo = [(200, 0, 0), (0, 120, 220), (0, 150, 80), (180, 90, 0), (140, 0, 180)]
        for cycle in sorted(ruta_picking["cycle"].unique().tolist()):
            ciclo_df = ruta_picking[ruta_picking["cycle"] == cycle].sort_values("orden_en_ciclo")
            puntos = [(dock_px, dock_py)]
            for _, row in ciclo_df.iterrows():
                px = margin + row["warehouse_x"] * scale
                py = margin + row["warehouse_y"] * scale
                puntos.append((px, py))
            color = colores_ciclo[(cycle - 1) % len(colores_ciclo)]
            if len(puntos) > 1:
                draw.line(puntos, fill=color, width=3)
            for _, row in ciclo_df.iterrows():
                px = margin + row["warehouse_x"] * scale
                py = margin + row["warehouse_y"] * scale
                draw.ellipse([px - 4, py - 4, px + 4, py + 4], fill=(0, 0, 0))
                draw.text((px + 5, py - 5), f"{int(row['cycle'])}.{int(row['orden_en_ciclo'])}", fill="black", font=font)

    leyenda_x = int(modelo.max_col * scale + margin + 25)
    draw.text((leyenda_x, margin), "Ruta de picking", fill="black", font=font)
    if not ruta_picking.empty:
        total = ruta_picking["distancia_desde_anterior"].sum()
        draw.text((leyenda_x, margin + 20), f"distancia aprox: {total:.1f}", fill="black", font=font)
        y = margin + 45
        for _, row in ruta_picking.head(16).iterrows():
            texto = f"{int(row['cycle'])}.{int(row['orden_en_ciclo'])} {row['ubicacion_almacen']} -> {row['slot_id']}"
            draw.text((leyenda_x, y), texto, fill="black", font=font)
            y += 16

    Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)
    img.save(ruta_salida)
    return str(Path(ruta_salida).resolve())


def optimizar_carga_camion(
    items: Sequence[ItemCarga],
    num_palets: int,
    *,
    pesos_objetivo: Optional[Dict[str, float]] = None,
) -> Dict[str, pd.DataFrame]:
    if not items:
        raise ValueError("La lista de items esta vacia.")

    pesos = {
        "acceso": 3.5,
        "almacen": 1.5,
        "espacio": 2.5,
        "mezcla_paradas": 2.0,
        "peso_lateral": 3.0,
        "peso_longitudinal": 2.0,
    }
    if pesos_objetivo:
        pesos.update(pesos_objetivo)

    total_paradas = max(item.stop_index for item in items)
    pico_retorno = _pico_neto_retorno(items, total_paradas)
    slots = _construir_slots_camion(num_palets, reserva_retorno_total_cm3=pico_retorno)

    entregas = [item for item in items if not item.es_retorno]
    retornos = [item for item in items if item.es_retorno]

    entregas_ordenadas = sorted(
        entregas,
        key=lambda item: (
            -item.stop_index,
            item.stack_rank,
            -item.peso_kg,
            -item.base_area_cm2,
            -item.alto_cm,
            item.warehouse_rank,
        ),
    )

    colocaciones: List[ColocacionEntrega] = []
    no_cargados: List[ItemCarga] = []

    for item in entregas_ordenadas:
        mejor_score = math.inf
        mejor_slot: Optional[SlotCamion] = None
        for slot in slots:
            slot_prueba = copy.deepcopy(slot)
            colocacion_prueba = slot_prueba.intentar_colocar_entrega(item)
            if not colocacion_prueba:
                continue
            slots_referencia = [slot_prueba if s.slot_id == slot.slot_id else s for s in slots]
            score = _score_slot_entrega(slot_prueba, item, slots_referencia, pesos, total_paradas)
            if score < mejor_score:
                mejor_score = score
                mejor_slot = slot

        if mejor_slot is None:
            no_cargados.append(item)
            continue

        colocacion_real = mejor_slot.intentar_colocar_entrega(item)
        if colocacion_real is None:
            no_cargados.append(item)
            continue
        colocaciones.append(colocacion_real)

    plan_retorno: List[PlanRetorno] = []
    retornos_no_asignados: List[ItemCarga] = []

    retornos_ordenados = sorted(
        retornos,
        key=lambda item: (item.stop_index, item.stack_rank, -item.base_area_cm2, -item.peso_kg),
    )
    for item in retornos_ordenados:
        mejor_score = math.inf
        mejor_slot = None
        for slot in slots:
            score = _score_slot_retorno(slot, item, item.stop_index, slots)
            if score < mejor_score:
                mejor_score = score
                mejor_slot = slot

        if mejor_slot is None or mejor_score == math.inf:
            retornos_no_asignados.append(item)
            continue

        estrategia = "hueco_liberado"
        if sum(vol for stop, vol in mejor_slot.entregas_por_stop.items() if stop <= item.stop_index) < item.volumen_cm3:
            estrategia = "reserva_superior"

        mejor_slot.volumen_retorno_asignado_cm3 += item.volumen_cm3
        plan_retorno.append(
            PlanRetorno(
                uid=item.uid,
                material=item.material,
                descripcion=item.descripcion,
                stop_id=item.stop_id,
                stop_name=item.stop_name,
                stop_index=item.stop_index,
                slot_id=mejor_slot.slot_id,
                lado=mejor_slot.lado,
                fila=mejor_slot.fila,
                peso_kg=item.peso_kg,
                volumen_cm3=item.volumen_cm3,
                estrategia=estrategia,
            )
        )

    df_colocaciones = pd.DataFrame([vars(c) for c in colocaciones])
    df_retorno = pd.DataFrame([vars(r) for r in plan_retorno])
    df_no_cargados = pd.DataFrame(
        [
            {
                "uid": item.uid,
                "material": item.material,
                "descripcion": item.descripcion,
                "stop_name": item.stop_name,
                "stop_index": item.stop_index,
                "es_retorno": item.es_retorno,
                "volumen_cm3": item.volumen_cm3,
                "peso_kg": item.peso_kg,
                "ubicacion_almacen": item.ubicacion_almacen,
                "warehouse_component": item.warehouse_component,
            }
            for item in (no_cargados + retornos_no_asignados)
        ]
    )

    resumen_slots = []
    for slot in slots:
        resumen_slots.append(
            {
                "slot_id": slot.slot_id,
                "lado": slot.lado,
                "fila": slot.fila,
                "alto_max_entrega_cm": slot.alto_max_entrega_cm,
                "altura_usada_cm": slot.altura_usada_cm(),
                "volumen_entregado_cm3": slot.volumen_entregado_cm3,
                "peso_entregado_kg": slot.peso_entregado_kg,
                "reserva_retorno_cm3": slot.reserva_retorno_cm3,
                "volumen_retorno_asignado_cm3": slot.volumen_retorno_asignado_cm3,
                "stops_distintos": len(slot.stops),
                "zonas_distintas": len(slot.zonas),
            }
        )
    df_slots = pd.DataFrame(resumen_slots).sort_values(["fila", "lado"])

    total_peso = df_slots["peso_entregado_kg"].sum() if not df_slots.empty else 0.0
    peso_izq = df_slots.loc[df_slots["lado"] == "L", "peso_entregado_kg"].sum() if not df_slots.empty else 0.0
    peso_der = df_slots.loc[df_slots["lado"] == "R", "peso_entregado_kg"].sum() if not df_slots.empty else 0.0
    utilizacion = df_slots["volumen_entregado_cm3"].sum() / max(1.0, sum(slot.volumen_total_disponible_cm3() for slot in slots))
    picking_por_ciclos = _construir_picking_por_ciclos(items, df_colocaciones)
    distancia_picking = picking_por_ciclos["distancia_desde_anterior"].sum() if not picking_por_ciclos.empty else 0.0
    palets_finales = _construir_palets_finales(items, df_colocaciones)
    factor_estiba_viaje = calcular_factor_estiba_viaje(items, num_palets)
    factible_total = len(no_cargados) == 0 and len(retornos_no_asignados) == 0

    metricas = pd.DataFrame(
        [
            {
                "palets": num_palets,
                "items_entrega_cargados": len(df_colocaciones),
                "items_retorno_planificados": len(df_retorno),
                "items_fuera_plan": len(df_no_cargados),
                "utilizacion_entrega": utilizacion,
                "peso_total_kg": total_peso,
                "desbalance_lateral_pct": abs(peso_izq - peso_der) / max(1.0, peso_izq + peso_der),
                "pico_retorno_reservado_cm3": pico_retorno,
                "distancia_picking_aprox": distancia_picking,
                "factible_total": factible_total,
            }
        ]
    )

    return {
        "colocaciones_entrega": df_colocaciones.sort_values(["stop_index", "fila", "layer_index", "y_cm", "x_cm"])
        if not df_colocaciones.empty
        else df_colocaciones,
        "plan_retorno": df_retorno.sort_values(["stop_index", "fila", "slot_id"]) if not df_retorno.empty else df_retorno,
        "resumen_slots": df_slots,
        "items_fuera_plan": df_no_cargados,
        "metricas": metricas,
        "picking_por_ciclos": picking_por_ciclos,
        "palets_finales": palets_finales,
        "factor_estiba_viaje": factor_estiba_viaje,
    }


def ejemplo_uso() -> None:
    items, _, _ = cargar_items_desde_excels(
        ruta_hackaton="Hackaton.xlsx",
        ruta_layout="Layout Mollet.xlsx",
        ruta="DR0001",
        max_stops=5,
    )
    resultado = optimizar_carga_camion(items, num_palets=6)
    print(resultado["metricas"].to_string(index=False))
    print(resultado["picking_por_ciclos"].head(20).to_string(index=False))
    print(resultado["palets_finales"].head(20).to_string(index=False))
    print(resultado["factor_estiba_viaje"].to_string(index=False))


def main() -> None:
    parser = argparse.ArgumentParser(description="Optimizador heuristico de carga por palets.")
    parser.add_argument("--hackaton", default="Hackaton.xlsx", help="Ruta al fichero Hackaton.xlsx")
    parser.add_argument("--zm040", help="Ruta opcional al fichero o CSV de catalogo ZM040. Si no se indica, usa el catalogo fijo del proyecto.")
    parser.add_argument("--layout", default="Layout Mollet.xlsx", help="Ruta al layout del almacen")
    parser.add_argument("--ruta", help="Codigo de ruta a analizar")
    parser.add_argument("--transporte", help="Codigo de transporte a analizar")
    parser.add_argument("--palets", type=int, default=6, choices=[3, 6, 8], help="Numero de huecos de palet del camion")
    parser.add_argument("--max-stops", type=int, help="Limita la prueba a las primeras N paradas de la ruta")
    parser.add_argument("--salida-prefix", help="Prefijo para exportar CSVs de salida")
    args = parser.parse_args()

    items, _, _ = cargar_items_desde_excels(
        ruta_hackaton=args.hackaton,
        ruta_zm040=args.zm040,
        ruta_layout=args.layout,
        ruta=args.ruta,
        transporte=args.transporte,
        max_stops=args.max_stops,
    )
    resultado = optimizar_carga_camion(items, num_palets=args.palets)

    print(resultado["metricas"].to_string(index=False))
    print(resultado["resumen_slots"].to_string(index=False))

    if args.salida_prefix:
        resultado["picking_por_ciclos"].to_csv(f"{args.salida_prefix}_picking_por_ciclos.csv", index=False)
        resultado["palets_finales"].to_csv(f"{args.salida_prefix}_palets_finales.csv", index=False)
        resultado["factor_estiba_viaje"].to_csv(f"{args.salida_prefix}_factor_estiba_viaje.csv", index=False)
        renderizar_plan_camion(resultado, f"{args.salida_prefix}_camion.png")
        renderizar_layout_picking(items, args.layout, resultado, f"{args.salida_prefix}_almacen.png")


if __name__ == "__main__":
    main()
